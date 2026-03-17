#!/usr/bin/env python3

import argparse
import os
import io
import re
import sys
from datetime import date, datetime
import requests
import csv
from openpyxl import Workbook, load_workbook
from typing import Any, Optional, Tuple
from configparser import ConfigParser
from pathlib import Path


# Mapping terms to GenCC IDs
allelic_requirement = {
    "biallelic_autosomal": "HP:0000007",
    "monoallelic_autosomal": "HP:0000006",
    "monoallelic_X_hemizygous": "HP:0001417",
    "monoallelic_Y_hemizygous": "HP:0001450",
    "monoallelic_X_heterozygous": "HP:0001417",
    "mitochondrial": "HP:0001427",
    "monoallelic_PAR": "HP:0000006",
    "biallelic_PAR": "HP:0000007",
}
confidence_category = {
    "definitive": "GENCC:100001",
    "strong": "GENCC:100002",
    "moderate": "GENCC:100003",
    "limited": "GENCC:100004",
    "disputed": "GENCC:100005",
    "refuted": "GENCC:100006",
}
submitter_id = "GENCC:000112"
submitter_name = "G2P"


def fetch_g2p_records(data: dict[str, Any]) -> list:
    """
    Fetches the G2P record using download for all panels.
    User is not authenticated to ensure only visible panels are downloaded

    Args:
        data (dict[str, Any]): The DB and API configuration dictionary

    Returns:
        list: The list reader
    """
    url = f"{data['api_url'].rstrip('/')}/panel/all/download/"
    response = requests.get(url)

    if response.status_code == 200:
        csv_content = io.StringIO(response.content.decode("utf-8"))
        reader = csv.DictReader(csv_content)
        return list(reader)

    else:
        sys.exit(f"Issues downloading the G2P file from {url}")


def write_to_the_GenCC_file(
    records_data: list,
    outfile: Path,
    output_file_issues: Path,
) -> Path:
    """
    Write the records to be submitted to the output file 'G2P_GenCC.txt'.
    Records without disease id are excluded from the submission and written
    to file 'record_with_issues.txt'.

    Args:
        records_data (dict[str, Any]): Record of unsubmitted ids
        outfile (Path): Output file to write the submission into
        output_file_issues (Path): File to write the records with issues into

    Returns:
        outfile (Path): The output file with records to be submitted
    """
    with open(outfile, mode="w") as rw:
        rw.write(
            "sgc_id\taction\tlocal_key\thgnc_id\thgnc_symbol\tdisease_id\tdisease_name\tmoi_id\tmoi_name"
            "\tsubmitter_id\tsubmitter_name\tclassification_id\tclassification_name\tdate"
            "\tpublic_report_url\tnotes\tpmids\tassertion_criteria_url\n"
        )
        issues_with_record = {}
        assertion_criteria_url = (
            "https://www.ebi.ac.uk/gene2phenotype/about/terminology"
        )

        for record in records_data:
            # Action 'U' (delete) should be written to the file first
            # The record object has different format from the other actions
            try:
                action = record["action"]
            except KeyError:
                action = "N"

            try:
                submission_id = record["sgc_id"]
            except KeyError:
                submission_id = ""

            if action == "U":
                line_to_output = f"{submission_id}\t{action}\n"
                rw.write(line_to_output)
                continue

            g2p_id = record["g2p id"]
            hgnc_id = record["hgnc id"]
            hgnc_symbol = record["gene symbol"]
            disease_id = record["disease mim"] or record["disease MONDO"]

            # We cannot submit records without disease ID
            if disease_id is None or disease_id == "":
                issues_with_record[g2p_id] = "Missing disease ID"
                continue

            disease_name = record["disease name"]

            if record["allelic requirement"] in allelic_requirement:
                moi_id = allelic_requirement[record["allelic requirement"]]
            else:
                issues_with_record[g2p_id] = (
                    f"Unsupported allelic requirement '{record['allelic requirement']}'"
                )
                continue

            moi_name = record["allelic requirement"]
            classification_id = confidence_category[record["confidence"]]
            classification_name = record["confidence"]
            dt = datetime.fromisoformat(record["date of last review"])
            date = dt.strftime("%Y/%m/%d")
            record_url = "https://www.ebi.ac.uk/gene2phenotype/lgd/" + g2p_id
            pmids = record["publications"]

            line_to_output = f"{submission_id or ''}\t{action}\t\t{hgnc_id}\t{hgnc_symbol}\t{disease_id}\t{disease_name}\t{moi_id}\t{moi_name}\t{submitter_id}\t{submitter_name}\t{classification_id}\t{classification_name}\t{date}\t{record_url}\t\t{pmids}\t{assertion_criteria_url}\n"
            rw.write(line_to_output)

    if issues_with_record:
        with open(output_file_issues, mode="a") as textfile:
            for g2p_id in issues_with_record:
                textfile.write(f"{g2p_id}\t{issues_with_record[g2p_id]}\n")

    return outfile


def create_datetime_now() -> date:
    """Creates datetime at present and formats it to YYYY-MM-DD

    Returns:
        date: Formatted date
    """

    time_now = datetime.now()
    db_date = time_now.strftime("%Y-%m-%d")
    return db_date


def convert_txt_to_excel(input_file: Path, output_file: Path):
    """Converts txt to Excel file

    Args:
        input_file (Path): Text file to be converted
        output_file (Path): Excel file that will be created
    """
    wb = Workbook()
    ws = wb.active

    delimiter = "\t"

    with open(input_file, "r") as file:
        for row_index, line in enumerate(file, start=1):
            for col_index, value in enumerate(line.rstrip("\n").split(delimiter), start=1):
                ws.cell(row=row_index, column=col_index, value=value)
    wb.save(output_file)


def read_from_config_file(config_file: str) -> dict[str, Any]:
    """
    Reads the database and api info from the config file.

    Args:
        config_file (str): configuration file

    Returns:
        dict[str, Any]: DB/API configuration dict
    """
    data = {}
    config = ConfigParser()
    config.read(config_file)
    data["api_url"] = config["api"]["api_url"]

    return data


def prepare_output_dir(path: str) -> Path:
    """
    Create the output directory where files are going to be written.

    Args:
        path (str): path where the G2P and GenCC files are going to be saved

    Returns:
        Path: output directory
    """
    timestamp = create_datetime_now()
    base_dir = Path(path) if path else Path.cwd()
    gencc_dir = base_dir / timestamp
    gencc_dir.mkdir(parents=True, exist_ok=True)

    return gencc_dir


def handle_existing_submission(
    latest_g2p_data: list,
    output_file: Path,
    output_file_updated: Path,
    output_file_issues: Path,
    output_file_deleted: Path,
    gencc_g2p_data: Path,
) -> Tuple[Path, Path, Path]:
    """
    Method to generate the files for a new submission to GenCC.
    Some records are new, others might have been updated since last submission
    and others might have been deleted in G2P.

    Args:
        latest_g2p_data (list): Data from the G2P all records file
        output_file (Path): Output text file where the submission will be written into
        output_file_updated (Path): Output text file where the updated records will be written into
        output_file_issues (Path): Output text file where the records with issues will be written into
        output_file_deleted (Path): Output text file where the records that have been deleted in G2P will be written into
        gencc_g2p_data (Path): Latest GenCC data
    """
    records_to_submit = []
    records_to_update = []
    records_to_delete = []
    records_with_issues = []
    g2p_ids_live = set()
    duplicated_records = {}

    # For each G2P record determine the action:
    #  - submit (N=new)
    #  - resubmit (R=republish)
    #  - delete (U=unpublish)
    #  - no action
    for g2p_record in latest_g2p_data:
        g2p_id = g2p_record["g2p id"]
        g2p_mondo_id = g2p_record["disease MONDO"].strip()
        g2p_omim_id = g2p_record["disease mim"].strip()

        # Keep track of unique records: gene + genotype + disease id (Mondo id) + mechanism
        key_by_mondo = f"{g2p_record['gene symbol']}---{g2p_mondo_id}---{g2p_record['allelic requirement']}---{g2p_record['molecular mechanism']}"

        if key_by_mondo in duplicated_records and g2p_mondo_id:
            records_with_issues.append(f"{g2p_id}\tfound duplicated record {duplicated_records[key_by_mondo]}")
        else:
            duplicated_records[key_by_mondo] = g2p_id

        # Save the G2P IDs to use later
        g2p_ids_live.add(g2p_id)

        # Action: submit (N=new)
        if g2p_id not in gencc_g2p_data.keys():
            # We can only submit if the record has Mondo or OMIM ID
            # This check is done by method write_to_the_GenCC_file()
            g2p_record["action"] = "N"
            g2p_record["sgc_id"] = None
            records_to_submit.append(g2p_record)
        else:
            gencc_row = gencc_g2p_data[g2p_id]
            old_confidence = gencc_row["classification_title"].lower()
            old_submitted_disease_id = gencc_row["submitted_as_disease_id"].replace(
                "OMIM:", ""
            )  # not sure we submitted this one
            old_mondo_id = gencc_row["disease_curie"]
            old_omim_id = gencc_row["disease_original_curie"].replace("OMIM:", "")

            # Action: resubmit (R=republish)
            if (
                (g2p_omim_id or g2p_mondo_id)
                and (
                    old_submitted_disease_id != g2p_omim_id
                    and old_submitted_disease_id != g2p_mondo_id
                )
            ) or old_confidence != g2p_record["confidence"]:
                g2p_record["action"] = "R"
                g2p_record["sgc_id"] = gencc_row["sgc_id"]
                records_to_update.append(g2p_record)

    for g2p_id in gencc_g2p_data:
        # Action: delete (U=unpublish)
        if g2p_id not in g2p_ids_live:
            gencc_record = gencc_g2p_data[g2p_id]
            gencc_record["action"] = "U"
            records_to_delete.append(gencc_record)

    # Write the new G2P records to submit to the GenCC file
    outfile = write_to_the_GenCC_file(
        records_to_submit, output_file, output_file_issues
    )

    # Write the updated records to a separate file
    outfile_updated_records = write_to_the_GenCC_file(
        records_to_update, output_file_updated, output_file_issues
    )

    outfile_deleted_records = write_to_the_GenCC_file(
        records_to_delete, output_file_deleted, output_file_issues
    )

    with open(output_file_issues, mode="a") as textfile:
        for issue_info in records_with_issues:
            textfile.write(issue_info+"\n")

    return outfile, outfile_updated_records, outfile_deleted_records


def clean_disease_name(text: str) -> Optional[str]:
    """
    Remove leading '<gene>-related' from the disease name.
    Called by: build_lgd_record_summary() in LocusGenotypeDiseaseSerializer
    """
    if not text:
        return None
    text = re.sub(r"^\S+-related\s+", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\(.*\)$", "", text)
    return text.strip() or None


def get_gencc_g2p_data(gencc_file):
    """
    Get all the G2P records from the GenCC file.
    These records are the reference data we use to determine records
    to submit/resubmit/delete.
    """
    gencc_g2p_data = {}

    workbook = load_workbook(gencc_file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]

    gencc_data = []
    for row in rows[1:]:
        row_dict = dict(zip(headers, row))
        gencc_data.append(row_dict)

    for row in gencc_data:
        if row["submitted_as_submitter_id"] == submitter_id:
            url = row["submitted_as_public_report_url"]
            # Get the G2P ID from the URL
            g2p_id = url.replace("https://www.ebi.ac.uk/gene2phenotype/lgd/", "")

            if g2p_id.startswith("G2P"):
                if g2p_id not in gencc_g2p_data:
                    gencc_g2p_data[g2p_id] = row

    return gencc_g2p_data


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "-p",
        "--path",
        help="Path where the G2P and GenCC files are going to be saved",
        required=False,
    )
    ap.add_argument("--config_file", required=True, help="G2P Configuration file")
    ap.add_argument(
        "--new_submission",
        action="store_true",
        required=False,
        help="Option to determine it is a new submission to GenCC",
    )
    ap.add_argument(
        "--gencc_file",
        required=True,
        type=Path,
        help="Latest GenCC data file",
    )
    args = ap.parse_args()

    if args.gencc_file and not os.path.isfile(args.gencc_file):
        sys.exit(f"Invalid file '{args.gencc_file}'")

    db_config = read_from_config_file(args.config_file)

    # Prepare output directory
    gencc_dir = prepare_output_dir(args.path)
    # File to save records that cannot be submitted due to issues
    output_file_issues = gencc_dir / "records_with_issues.txt"
    # File to save records that are going to be submitted
    output_file = gencc_dir / "G2P_GenCC.txt"
    # File to save records that are going to be submitted as updates
    output_file_updated = gencc_dir / "G2P_GenCC_updated_records.txt"
    # File to save records that have been deleted in G2P
    output_file_deleted = gencc_dir / "G2P_GenCC_deleted_records.txt"
    # File to save records that have been deleted in G2P (expected format for GenCC)
    final_output_file_deleted = gencc_dir / "G2P_GenCC_deleted_records.xlsx"
    # File to save records that are going to be submitted (expected format for GenCC)
    final_output_file = gencc_dir / "G2P_GenCC.xlsx"
    # File to save records that are going to be re-submitted (expected format for GenCC)
    final_output_file_updated = gencc_dir / "G2P_GenCC_updated_records.xlsx"

    print("Getting G2P records from the API...")
    g2p_data = fetch_g2p_records(db_config)
    print("Getting G2P records from the API... done")

    if args.new_submission:
        print("\nHandling new submission...")
        # A new submission means that all records are going to be submitted
        # for the first time
        outfile = write_to_the_GenCC_file(
            g2p_data, output_file, output_file_issues
        )
        print("Handling new submission... done")
    else:
        print("\nHandling existing submission...")
        print(f"Getting G2P data from GenCC file {args.gencc_file}")
        gencc_g2p_data = get_gencc_g2p_data(args.gencc_file)

        outfile, outfile_updated_records, outfile_deleted_records = handle_existing_submission(
            g2p_data,
            output_file,
            output_file_updated,
            output_file_issues,
            output_file_deleted,
            gencc_g2p_data,
        )

    print("\nConverting text file to Excel file...")
    convert_txt_to_excel(outfile, final_output_file)
    if not args.new_submission:
        convert_txt_to_excel(outfile_updated_records, final_output_file_updated)
        convert_txt_to_excel(outfile_deleted_records, final_output_file_deleted)
    print("Converting text file to Excel file... done")


if __name__ == "__main__":
    main()
