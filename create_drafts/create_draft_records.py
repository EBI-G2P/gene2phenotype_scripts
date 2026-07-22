#!/usr/bin/env python3

import sys
import argparse
import configparser
import csv
import os.path
import requests
import json


"""
Script to read the pre-draft data from ClinGen or PanelApp JSON file and create draft records.

Options:
        --source         ClinGen or PanelApp
        --input_file     Data to be used to create drafts (supported format: json)

Example how to run:
    python create_draft_records.py \
        --config config.ini \
        --source ClinGen \
        --input_file clingen_reviewed_data.json
"""

valid_mechanisms = [
    "loss of function",
    "gain of function",
    "dominant negative",
    "undetermined non loss of function",
    "undetermined",
]

allelic_requirement_mapping = {
    "autosomal recessive": "biallelic_autosomal",
    "biallelic": "biallelic_autosomal",
    "autosomal dominant": "monoallelic_autosomal",
    "X-linked": "monoallelic_X",
}

MAX_SESSION_NAME_LENGTH = 100


def build_session_name(*parts: str) -> str:
    """Build a DB-safe session name within the column length limit."""
    return "_".join(parts)[:MAX_SESSION_NAME_LENGTH]


def normalize_excel_headers(headers: tuple) -> list[str]:
    """Make worksheet headers usable as unique dictionary keys."""
    counts = {}
    normalized_headers = []

    for index, header in enumerate(headers, start=1):
        header_text = str(header).strip() if header is not None else f"column_{index}"
        if not header_text:
            header_text = f"column_{index}"

        counts[header_text] = counts.get(header_text, 0) + 1
        if counts[header_text] > 1:
            header_text = f"{header_text}_{counts[header_text]}"

        normalized_headers.append(header_text)

    return normalized_headers


def login(
    api_username: str, api_password: str, api_url: str
) -> requests.cookies.RequestsCookieJar:
    """Login into G2P API"""
    login_url = f"{api_url.rstrip('/')}/login/"

    response = requests.post(
        login_url, json={"username": api_username, "password": api_password}
    )

    if response.status_code != 200:
        sys.exit("Login failed. Check your credentials and API URL.")

    return response.cookies


def logout(api_url: str, cookies: requests.cookies.RequestsCookieJar) -> None:
    """Logout of the API"""
    logout_url = f"{api_url.rstrip('/')}/logout/"

    response = requests.post(logout_url, cookies=cookies)

    if response.status_code != 204:
        sys.exit("Logout failed. Check your credentials and API URL.")


def get_all_automatic_drafts(
    api_url: str, cookies: requests.cookies.RequestsCookieJar
) -> dict:
    """Fetch all automatic draft records from G2P"""
    drafts = {}

    url = f"{api_url.rstrip('/')}/curations/?scope=all&type=automatic"
    response = requests.get(url, cookies=cookies)
    if response.status_code != 200:
        sys.exit(f"Failed to fetch draft records: {response.text}")
    else:
        data_response = response.json()["results"]
        for draft in data_response:
            drafts[draft["session_name"]] = draft

    return drafts


def insert_draft_record(
    api_url: str, cookies: requests.cookies.RequestsCookieJar, draft_data: dict
) -> None:
    """
    Insert a draft record into G2P
    The endpoint accepts the following data format:
        {
            "json_data": {...},
            "status": "automatic"  # optional, default is 'manual'
        }
    """
    url = f"{api_url.rstrip('/')}/add/curation/"
    response = requests.post(url, json=draft_data, cookies=cookies)
    if response.status_code != 200:
        sys.exit(f"Failed to insert draft record: {response.text}")
    else:
        print(f"Draft record inserted successfully: {response.json()}")


def fetch_pmid_info(api_url: str, pmid: str) -> dict:
    url = f"{api_url.rstrip('/')}/publication/{pmid}"
    response = requests.get(url)
    if response.status_code != 200:
        sys.exit(f"Cannot fetch PMID {pmid}")
    else:
        return response.json()["results"][0]


def fetch_disease_info(api_url: str, disease_id: str) -> dict:
    url = f"{api_url.rstrip('/')}/external_disease/{disease_id}"
    response = requests.get(url)
    if response.status_code != 200:
        sys.exit(f"Cannot fetch disease {disease_id}")
    else:
        return response.json()["results"][0]


def prepare_draft_records(
    input_file: str,
    panel_name: str,
    source: str,
    automatic_drafts: dict,
    api_url: str,
    cookies: requests.cookies.RequestsCookieJar,
) -> None:
    """
    Read the input json file
    """
    existing_records = []

    with open(input_file) as fh:
        data = json.load(fh)
        for record in data:
            final_draft = {}

            if (("status" in record and (record["status"] == "approved" or record["status"] == "needs changes"))
                or ("type" in record and record["type"] != "in_g2p")):
                if "comments" in record and "g2p_data" in record["comments"] and (
                    "matches the existing g2p" in record["comments"]["g2p_data"].lower()
                    or "draft matches" in record["comments"]["g2p_data"].lower()
                ):
                    existing_records.append(record)
                else:
                    final_draft["extra_comment"] = ""
                    # Add an extra comment related to the necessary changes
                    if "status" in record and record["status"] == "needs changes":
                        final_draft["extra_comment"] += (
                            "Note: this draft needs changes" + "\n"
                        )
                    if "comments" in record and "g2p_data" in record["comments"]:
                        final_draft["extra_comment"] += (
                            "Existing G2P data: "
                            + record["comments"]["g2p_data"]
                            + "\n"
                        )

                    final_draft["locus"] = record["gene_symbol"]
                    final_draft["panels"] = [panel_name]
                    final_draft["public_comment"] = ""
                    final_draft["private_comment"] = ""
                    final_draft["cross_cutting_modifier"] = []
                    final_draft["variant_types"] = []
                    final_draft["variant_descriptions"] = []
                    final_draft["variant_consequences"] = []

                    # Add source details to JSON
                    final_draft["source_data"] = {"name": source}
                    if "url" in record:
                        final_draft["source_data"]["url"] = record["url"]

                    final_draft["confidence"] = ""
                    # Save the confidence in the source field
                    if "confidence" in record and record["confidence"] != "":
                        final_draft["source_data"]["confidence"] = record[
                            "confidence"
                        ].lower()
                    if "comments" in record and "confidence" in record["comments"]:
                        final_draft["extra_comment"] += (
                            "Confidence comment: "
                            + record["comments"]["confidence"]
                            + "\n"
                        )

                    final_draft["publications"] = []
                    for pmid in record["pmids"]:
                        publication_data = fetch_pmid_info(api_url, pmid)
                        final_draft["publications"].append(
                            {
                                "pmid": pmid,
                                "year": publication_data["year"],
                                "title": publication_data["title"],
                                "source": publication_data["source"],
                                "authors": publication_data["authors"],
                                "comment": "",
                                "families": None,
                                "ancestries": "",
                                "consanguineous": "unknown",
                                "affectedIndividuals": None,
                            }
                        )

                    final_draft["phenotypes"] = []
                    if record["phenotypes"]:
                        final_draft["source_data"]["phenotypes"] = record["phenotypes"]

                    # Format allelic requirement
                    final_draft["allelic_requirement"] = ""
                    if (
                        record["allelic_requirement"]
                        in allelic_requirement_mapping.keys()
                    ):
                        final_draft["allelic_requirement"] = (
                            allelic_requirement_mapping[record["allelic_requirement"]]
                        )
                    else:
                        final_draft["extra_comment"] += (
                            "Unsupported allelic requirement: "
                            + record["allelic_requirement"]
                            + "\n"
                        )

                    # Mechanism
                    final_draft["source_data"]["mechanism_comment"] = ""
                    # Check if the mechanism value is valid, if not add it to the extra_comment
                    valid_mechanism = ""
                    if (
                        "mechanism" in record and record["mechanism"] != ""
                        and record["mechanism"].lower().replace("-", " ")
                        not in valid_mechanisms
                    ):
                        final_draft["source_data"]["mechanism_comment"] += (
                            "Unsupported mechanism: " + record["mechanism"]
                        )
                    else:
                        if "mechanism" in record and record["mechanism"] != "":
                            valid_mechanism = record["mechanism"].lower().replace("-", " ")

                    final_draft["molecular_mechanism"] = {
                        "name": "",
                        "support": "",
                    }
                    final_draft["mechanism_synopsis"] = []
                    # Mechanism evidence
                    final_draft["mechanism_evidence"] = []
                    # Save mechanism, mechanism evidence and comment in the source field
                    final_draft["source_data"]["mechanism"] = valid_mechanism
                    if "evidence" in record and record["evidence"]:
                        final_draft["source_data"]["mechanism_evidence"] = "; ".join(
                            record["evidence"]
                        )
                    if "comments" in record and "mechanism" in record["comments"]:
                        final_draft["source_data"]["mechanism_comment"] += (
                            "\n" + record["comments"]["mechanism"]
                        )

                    # Disease
                    disease_cross_references = []
                    if "mondo_id" in record and record["mondo_id"] != "":
                        disease_data = fetch_disease_info(api_url, record["mondo_id"])
                        disease_cross_references.append(
                            {
                                "source": "Mondo",
                                "identifier": record["mondo_id"],
                                "disease_name": disease_data["disease"].lower(),
                                "original_disease_name": disease_data["disease"],
                            }
                        )
                    if "disease_id" in record and record["disease_id"] != "" and record["disease_id"] != []:
                        omim_list = record["disease_id"].split(",")
                        for omim_id in omim_list:
                            omim_id = omim_id.replace("OMIM:", "")
                            omim_id = omim_id.replace("MIM:", "")
                            if omim_id.strip().isdigit():
                                disease_data = fetch_disease_info(
                                    api_url, omim_id.strip()
                                )
                                disease_cross_references.append(
                                    {
                                        "source": "OMIM",
                                        "identifier": omim_id.strip(),
                                        "disease_name": disease_data["disease"].lower(),
                                        "original_disease_name": disease_data[
                                            "disease"
                                        ],
                                    }
                                )
                    final_draft["disease"] = {
                        "disease_name": "",
                        "cross_references": [],
                    }
                    # Save the disease name and cross references in the source field
                    if "disease" in record:
                        final_draft["source_data"]["disease"] = record["disease"]
                    final_draft["source_data"]["disease_cross_references"] = (
                        disease_cross_references
                    )

                    # Build the session name using gene, allelic requirement and disease to check for duplicates before inserting the draft
                    final_draft["session_name"] = build_session_name(
                        record["gene_symbol"],
                        final_draft["allelic_requirement"],
                        record["disease"],
                    )

                    # Check if the draft already exists based on the session name
                    if final_draft["session_name"] in automatic_drafts:
                        print(
                            f"Draft record for session '{final_draft['session_name']}' already exists. Skipping insertion."
                        )
                        continue

                    # print("\n\n->", final_draft)

                    # Call G2P API to insert the curation draft
                    insert_draft_record(
                        api_url,
                        cookies,
                        {"json_data": final_draft, "status": "automatic"},
                    )


def prepare_draft_records_pipeline(
    input_file: str,
    g2p_to_pmids: dict,
    panel_name: str,
    source: str,
    automatic_drafts: dict,
    api_url: str,
    cookies: requests.cookies.RequestsCookieJar,
) -> None:
    """
    Read the input csv file. This is the fake G2P file that was used by the pipeline.
    """

    with open(input_file) as fh:
        data = csv.DictReader(fh)
        for record in data:
            final_draft = {}

            final_draft["locus"] = record["gene symbol"]
            final_draft["panels"] = [panel_name]
            final_draft["public_comment"] = ""
            final_draft["private_comment"] = ""
            final_draft["cross_cutting_modifier"] = []
            final_draft["variant_types"] = []
            final_draft["variant_descriptions"] = []
            final_draft["variant_consequences"] = []
            final_draft["confidence"] = ""
            final_draft["phenotypes"] = []
            final_draft["allelic_requirement"] = record["allelic requirement"]

            # Publications
            final_draft["publications"] = []
            if record["g2p id"] in g2p_to_pmids:
                for pmid in g2p_to_pmids[record["g2p id"]]:
                    publication_data = fetch_pmid_info(api_url, pmid)
                    final_draft["publications"].append(
                        {
                            "pmid": pmid,
                            "year": publication_data["year"],
                            "title": publication_data["title"],
                            "source": publication_data["source"],
                            "authors": publication_data["authors"],
                            "comment": "",
                            "families": None,
                            "ancestries": "",
                            "consanguineous": "unknown",
                            "affectedIndividuals": None,
                        }
                    )

            # Mechanism
            final_draft["molecular_mechanism"] = {
                "name": "",
                "support": "",
            }
            final_draft["mechanism_synopsis"] = []
            # Mechanism evidence
            final_draft["mechanism_evidence"] = []

            # Disease
            final_draft["disease"] = {
                "disease_name": "dilated cardiomyopathy",
                "cross_references": [],
            }

            # Build the session name using gene, allelic requirement and disease to check for duplicates before inserting the draft
            final_draft["session_name"] = build_session_name(
                record["gene symbol"],
                record["allelic requirement"],
                final_draft["disease"]["disease_name"],
            )

            # Check if the draft already exists based on the session name
            if final_draft["session_name"] in automatic_drafts:
                print(
                    f"Draft record for session '{final_draft['session_name']}' already exists. Skipping insertion."
                )
                continue

            final_draft["source_data"] = {"name": source}

            if not final_draft["publications"]:
                print(f"No publications found for {record['gene symbol']}.")
                continue

            # ### TO REMOVE ###
            # if record['gene symbol'] == "CRYAB":
            #     continue

            print(f"Creating draft for {record['gene symbol']}")

            if (record['gene symbol'] == "EPG5" or record['gene symbol'] == "PLN"
                or record['gene symbol'] == "ACTN2" or record['gene symbol'] == "TAFAZZIN"):
                print(f"Skipping draft creation for {record['gene symbol']}. Gene already in G2P.")
                continue

            # Call G2P API to insert the curation draft
            insert_draft_record(
                api_url,
                cookies,
                {"json_data": final_draft, "status": "automatic"},
            )


def prepare_publications_pipeline(pipeline_file):
    """
    """
    g2p_to_pmids = {}

    with open(pipeline_file) as fh:
        data = csv.DictReader(fh)
        for record in data:
            pmid = record["PMID"]
            g2p_id = record["G2P_IDs"]

            if "gemini_relevance_label" in record:
                score = record["gemini_relevance_label"]

                if score != "high":
                    continue

            if g2p_id not in g2p_to_pmids:
                g2p_to_pmids[g2p_id] = []
            g2p_to_pmids[g2p_id].append(pmid)

    return g2p_to_pmids


def prepare_draft_records_immuno(
    input_file: str, # clingen pre-draft data
    immuno_file: str, # immuno panel genes + other data
    panel_name: str,
    source: str,
    automatic_drafts: dict,
    api_url: str,
    cookies: requests.cookies.RequestsCookieJar,
) -> None:
    """
    Read the input file from the immuno panel to build the drafts records.
    It also attaches the clingen data to the draft record if the gene is present in the clingen pre-draft data.
    """
    import openpyxl

    ar_mapping = {
        "AR": "biallelic_autosomal",
        "AD": "monoallelic_autosomal",
        "XL": "monoallelic_X",
    }

    mechanism_mapping = {
        "GOF": "gain of function",
        "LOF": "loss of function",
        "DN": "dominant negative",
    }

    if not os.path.isfile(immuno_file):
        sys.exit(f"Invalid immuno file '{immuno_file}'")

    workbook = openpyxl.load_workbook(immuno_file, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        sys.exit(f"Immuno file '{immuno_file}' is empty")

    headers = normalize_excel_headers(raw_headers)
    immuno_records = []

    for row in rows:
        if row is None or not any(value is not None and str(value).strip() != "" for value in row):
            continue

        record = {}
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else None
            if isinstance(value, str):
                value = value.strip()
            record[header] = value

        # print("\n\n->", record)

        if "G2P record" in record and record["G2P record"] is not None:
            print(f"Gene is already in G2P: {record['Genetic defect']}. Skipping draft creation.")
            continue

        disease_name = record["Disease"].strip()
        gene_symbol = record["Genetic defect"].strip()

        if record["Inheritance"] is None:
            allelic_requirement = ""
        else:
            allelic_requirement = ar_mapping.get(record["Inheritance"].strip(), "")

        if record["GOF/DN"] is None:
            mechanism = ""
        else:
            mechanism = mechanism_mapping.get(record["GOF/DN"].strip(), "")

        # Check if mechanism is in disease name
        if mechanism == "":
            if "(LOF)" in disease_name or " LOF" in disease_name:
                mechanism = "loss of function"
            elif "(GOF)" in disease_name or " GOF" in disease_name:
                mechanism = "gain of function"
            elif "(DN)" in disease_name or " DN" in disease_name:
                mechanism = "dominant negative"

        # Attach data to record
        record["allelic_requirement"] = allelic_requirement
        record["mechanism"] = mechanism
        record["gene_symbol"] = gene_symbol
        record["disease_name"] = disease_name

        # print("---> Gene:", gene_symbol, "Disease:", disease_name, "AR:", allelic_requirement, "Mechanism:", mechanism)

        immuno_records.append(record)

    workbook.close()

    # Fetch the pre-draft data from ClinGen
    pre_draft_dict = {}
    with open(input_file) as fh:
        pre_draft_data = json.load(fh)
        # Build dict for clingen pre-draft data for faster lookup
        for record in pre_draft_data:
            if record["gene_symbol"] not in pre_draft_dict:
                pre_draft_dict[record["gene_symbol"]] = [record]
            else:
                pre_draft_dict[record["gene_symbol"]].append(record)

    # Start building the draft records using the pre-draft data and the immuno records
    # Keep track of gene symbols
    count_gene_symbols = {}

    for immuno_record in immuno_records:
        if immuno_record["gene_symbol"] not in pre_draft_dict:
            continue  # Skip if gene is not in the pre-draft data
        
        if immuno_record["gene_symbol"] in count_gene_symbols:
            count_gene_symbols[immuno_record["gene_symbol"]] += 1
        else:
            count_gene_symbols[immuno_record["gene_symbol"]] = 1

        disease_name = immuno_record["disease_name"]
        mechanism = immuno_record["mechanism"]

        pre_drafts_clingen = pre_draft_dict[immuno_record["gene_symbol"]]
        clingen_record = None

        # If there are several pre-drafts for the same gene, get the one with same allelic requirement
        # and mechanism. If not found, skip the gene and print a warning.
        if len(pre_drafts_clingen) > 1:
            for pre_draft in pre_drafts_clingen:
                ar = allelic_requirement_mapping.get(pre_draft["allelic_requirement"], "")
                if ar == immuno_record["allelic_requirement"] and (pre_draft["mechanism"] == mechanism or pre_draft["mechanism"] == "" or mechanism == ""):
                    clingen_record = pre_draft
                    break
        else:
            if count_gene_symbols[immuno_record["gene_symbol"]] > 1:
                print(f"WARNING: Multiple immuno rows for gene: {immuno_record['gene_symbol']}")
            
            # print("\n---> Gene (immuno file):", immuno_record["gene_symbol"], "AR:", immuno_record["allelic_requirement"])
            # print("---> Gene (clingen):", pre_drafts_clingen[0]["gene_symbol"], "AR:", pre_drafts_clingen[0]["allelic_requirement"])

            ar = allelic_requirement_mapping.get(pre_drafts_clingen[0]["allelic_requirement"], "")
            if ar == immuno_record["allelic_requirement"] and (pre_drafts_clingen[0]["mechanism"] == mechanism or pre_drafts_clingen[0]["mechanism"] == "" or mechanism == ""):
                clingen_record = pre_drafts_clingen[0]
            else:
                print(f"WARNING: No matching allelic requirement/mechanism for gene: {immuno_record['gene_symbol']}. Skipping.")
                continue

        final_draft = {}
        final_draft["locus"] = immuno_record["gene_symbol"]
        final_draft["panels"] = [panel_name]
        final_draft["public_comment"] = ""
        final_draft["private_comment"] = ""
        final_draft["cross_cutting_modifier"] = []
        final_draft["variant_types"] = []
        final_draft["variant_descriptions"] = []
        final_draft["variant_consequences"] = []
        final_draft["confidence"] = ""
        final_draft["phenotypes"] = []
        final_draft["allelic_requirement"] = immuno_record["allelic_requirement"]

        if clingen_record is None:
            print(f"WARNING: No matching ClinGen record found for gene: {immuno_record['gene_symbol']}, allelic requirement: {immuno_record['allelic_requirement']}, mechanism: {mechanism}. Skipping.")
            continue

        final_draft["publications"] = []
        for pmid in clingen_record["pmids"]:
            publication_data = fetch_pmid_info(api_url, pmid)
            if publication_data["authors"] is not None:
                final_draft["publications"].append(
                    {
                        "pmid": pmid,
                        "year": publication_data["year"],
                        "title": publication_data["title"],
                        "source": publication_data["source"],
                        "authors": publication_data["authors"],
                        "comment": "",
                        "families": None,
                        "ancestries": "",
                        "consanguineous": "unknown",
                        "affectedIndividuals": None,
                    }
                )

        final_draft["molecular_mechanism"] = {
            "name": mechanism,
            "support": "",
        }
        final_draft["mechanism_synopsis"] = []
        final_draft["mechanism_evidence"] = []

        final_draft["disease"] = {
            "disease_name": immuno_record["disease_name"],
            "cross_references": [],
        }

        final_draft["source_data"] = {"name": source}
        if "url" in clingen_record:
            final_draft["source_data"]["url"] = clingen_record["url"]

        final_draft["source_data"]["mechanism"] = clingen_record["mechanism"]
        if "evidence" in clingen_record and clingen_record["evidence"]:
            final_draft["source_data"]["mechanism_evidence"] = "; ".join(
                clingen_record["evidence"]
            )

        final_draft["source_data"]["phenotypes"] = clingen_record["phenotypes"]

        # Disease
        disease_cross_references = []
        if "mondo_id" in clingen_record and clingen_record["mondo_id"] != "":
            disease_data = fetch_disease_info(api_url, clingen_record["mondo_id"])
            disease_cross_references.append(
                {
                    "source": "Mondo",
                    "identifier": clingen_record["mondo_id"],
                    "disease_name": disease_data["disease"].lower(),
                    "original_disease_name": disease_data["disease"],
                }
            )
        if "disease_id" in clingen_record and clingen_record["disease_id"] != "" and clingen_record["disease_id"] != []:
            omim_list = clingen_record["disease_id"].split(",")
            for omim_id in omim_list:
                if omim_id.startswith("OMIM:") or omim_id.startswith("MIM:"):
                    omim_id = omim_id.replace("OMIM:", "")
                    omim_id = omim_id.replace("MIM:", "")
                    if omim_id.strip().isdigit():
                        disease_data = fetch_disease_info(
                            api_url, omim_id.strip()
                        )
                        disease_cross_references.append(
                            {
                                "source": "OMIM",
                                "identifier": omim_id.strip(),
                                "disease_name": disease_data["disease"].lower(),
                                "original_disease_name": disease_data[
                                    "disease"
                                ],
                            }
                        )
                elif omim_id.startswith("MONDO:"):
                    mondo_id = omim_id.strip()
                    disease_data = fetch_disease_info(
                        api_url, mondo_id
                    )
                    disease_cross_references.append(
                        {
                            "source": "Mondo",
                            "identifier": mondo_id,
                            "disease_name": disease_data["disease"].lower(),
                            "original_disease_name": disease_data[
                                "disease"
                            ],
                        }
                    )

        # Save the disease name and cross references in the source field
        if "disease" in clingen_record:
            final_draft["source_data"]["disease"] = clingen_record["disease"]
        final_draft["source_data"]["disease_cross_references"] = (
            disease_cross_references
        )

        final_draft["session_name"] = build_session_name(
            immuno_record["gene_symbol"],
            immuno_record["allelic_requirement"],
            immuno_record["disease_name"],
        )

        # Check if the draft already exists based on the session name
        if final_draft["session_name"] in automatic_drafts:
            print(
                f"Draft record for session '{final_draft['session_name']}' already exists. Skipping insertion."
            )
            continue

        # print(f"Creating draft for {immuno_record['gene_symbol']} with disease {immuno_record['disease_name']}")

        # print(json.dumps(final_draft, indent=4))

        # Call G2P API to insert the curation draft
        insert_draft_record(
            api_url,
            cookies,
            {"json_data": final_draft, "status": "automatic"},
        )


def main():
    parser = argparse.ArgumentParser(description="Create draft records")
    parser.add_argument("--source", required=True, help="ClinGen or PanelApp")
    parser.add_argument(
        "--input_file",
        required=True,
        help="Pre-draft data to be used to create drafts (supported format: json)",
    )
    parser.add_argument(
        "--pipeline_file",
        required=False,
        help="The output file generated by the mined publication pipeline with G2P IDs and PMIDs. Supported format: csv",
    )
    parser.add_argument(
        "--immuno_file",
        required=False,
        help="The file containing list of genes for immuno panel. Supported format: xlsx",
    )
    parser.add_argument(
        "--config", required=True, help="Config file with details to G2P API"
    )
    parser.add_argument(
        "--panel", required=True, help="G2P panel name e.g. 'Ear disorders'"
    )
    parser.add_argument(
        "--output_file", required=False, help="Output file to save the final draft records"
    )
    args = parser.parse_args()

    input_file = args.input_file
    source = args.source
    panel_name = args.panel
    output_file = args.output_file
    pipeline_file = args.pipeline_file
    immuno_file = args.immuno_file

    if pipeline_file and immuno_file:
        sys.exit("ERROR: Cannot provide both --pipeline_file and --immuno_file options. Please choose one.")

    if not output_file:
        output_file = f"{source}_draft_records_created.json"

    # Load the config file
    config = configparser.ConfigParser()
    config.read(args.config)

    try:
        api = config["api"]
    except KeyError:
        sys.exit("ERROR: 'api' missing from config file")
    else:
        api_url = api["api_url"]
        api_username = api["api_username"]
        api_password = api["api_password"]

    if not os.path.isfile(input_file):
        sys.exit(f"Invalid input file '{input_file}'")

    cookies = login(api_username, api_password, api_url)

    # Get all existing automatic drafts to check for duplicates before inserting new ones
    automatic_drafts = get_all_automatic_drafts(api_url, cookies)

    if pipeline_file:
        # Prepare the publications
        # Read the output file generated by the pipeline
        print(f"Preparing drafts from pipeline file: {pipeline_file}...")
        g2p_to_pmids = prepare_publications_pipeline(pipeline_file)

        prepare_draft_records_pipeline(
            input_file, g2p_to_pmids, panel_name, source.lower(), automatic_drafts, api_url, cookies
        )
        print(f"Preparing drafts from pipeline file: {pipeline_file}... done")
    elif immuno_file:
        # Prepare the immuno panel genes
        print(f"Preparing drafts from immuno file: {immuno_file}...")
        prepare_draft_records_immuno(
            input_file, immuno_file, panel_name, source.lower(), automatic_drafts, api_url, cookies
        )
        print(f"Preparing drafts from immuno file: {immuno_file}... done")
    else:
        print(f"Preparing drafts from source: {source} and input file: {input_file}...")
        prepare_draft_records(
            input_file, panel_name, source.lower(), automatic_drafts, api_url, cookies
        )
        print(f"Preparing drafts from source: {source} and input file: {input_file}... done")

    logout(api_url, cookies)


if __name__ == "__main__":
    main()
